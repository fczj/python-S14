import  re
import openpyxl
import sys


def read_file(f_name):
    with open(f_name,'r',encoding='utf8') as f:
        return f.read()


def find_item(text):
    name = re.compile('<tr><td height="80"><div class="cell-box ft-center">.*?<a href="http://.*?title="(.*?)">')
    zhanggui = re.compile(r'掌柜:</label>(.*?)</div>',re.S)
    qita = re.compile(r'"cell-box fm-arial ft-right">(.*?)</div></td><td>',re.S)
    names = (re.findall(name, text))
    zhanguis = (re.findall(zhanggui, text))
    qitas = (re.findall(qita,text))
    j = 0
    k = 5
    for i in range(10):
        ret = []
        ret.append(names[i])
        ret.append(zhanguis[i])
        ret.extend(qitas[j:k])
        j += 5
        k += 5
        yield  (ret)

filepath = "./test101.xlsx"
#wb = openpyxl.Workbook()
#wb.save(filepath)

wb = openpyxl.load_workbook(filename=filepath)
ws = wb.get_sheet_by_name('Sheet')
nowrow = ws.max_row + 1


for page in range(1):
    text = read_file(r'/home/lanjing/data/00{}.html'.format(int(page)+1))
    #text = read_file(r'/home/lanjing/data/010.html')
    for i in find_item(text):
        for ii,jj in enumerate(i):
            print (ii,jj)
            ws.cell(row=nowrow, column=ii + 1).value = jj
        nowrow +=1

wb.save(filename='{}'.format(filepath))
