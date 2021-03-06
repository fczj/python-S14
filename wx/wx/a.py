#!/usr/bin/env python
# -*- coding: utf-8 -*-
# @Time    : 17-8-27 下午12:23
# @Author  : xiongzhibiao
# @Email   : 158349411@qq.com
# @File    : 爬虫.py
# @Software: PyCharm


import requests
import re
import openpyxl
import sys
import os

LST = [
'https://detail.tmall.com/item.htm?id=520081219835',
'https://detail.tmall.com/item.htm?id=539985337215',
'https://detail.tmall.com/item.htm?id=521241458877',
'https://detail.tmall.com/item.htm?id=552884270113',
'https://detail.tmall.com/item.htm?id=549084199291',
'https://detail.tmall.com/item.htm?id=15171590054',
'https://detail.tmall.com/item.htm?id=38623634926',
'https://detail.tmall.com/item.htm?id=543366678573',
'https://detail.tmall.com/item.htm?id=524405966399',
'http://item.taobao.com/item.htm?id=544555172559',
'https://detail.tmall.com/item.htm?id=550333369029',
'https://detail.tmall.hk/hk/item.htm?id=550369625401',
'https://detail.tmall.hk/hk/item.htm?id=535004286373',
'https://detail.tmall.com/item.htm?id=532836939931&ali_refid=a3_419108_1006:1108596375:N:%E8%90%A5%E5%85%BB%E5%A5%B6%E6%98%94:0e4b8da32432d2c03c235a3fa25305b6&ali_trackid=1_0e4b8da32432d2c03c235a3fa25305b6&spm=a220m.1000858.0.0.4c637810ePmF1n',
'https://detail.tmall.com/item.htm?id=554869827041',
'https://detail.tmall.com/item.htm?id=547058495979',
'https://detail.tmall.com/item.htm?id=543869180545',
'https://detail.tmall.com/item.htm?id=553113341217',
'https://detail.tmall.hk/hk/item.htm?id=36252836776',
'https://detail.tmall.hk/hk/item.htm?id=525245020578',
'https://detail.tmall.com/item.htm?id=547580853094',
'https://detail.tmall.com/item.htm?id=19396701753',
'https://detail.tmall.com/item.htm?id=530258856669',
'https://detail.tmall.com/item.htm?id=549106402196',
'http://item.taobao.com/item.htm?id=552567844672',
'https://detail.tmall.com/item.htm?id=527294319332',
'https://detail.tmall.com/item.htm?id=553782404502',
'https://detail.tmall.com/item.htm?id=543441461751',
'https://detail.tmall.com/item.htm?id=543892457912',
'https://detail.tmall.com/item.htm?id=543891525043',
'https://detail.tmall.hk/hk/item.htm?id=40382917292',
'https://detail.tmall.com/item.htm?id=540619287982 ',
'https://detail.tmall.com/item.htm?id=539737608591',
'https://detail.tmall.com/item.htm?id=536232168801',
'https://detail.tmall.com/item.htm?id=543108313701',
'http://item.taobao.com/item.htm?id=38736330387',
'https://detail.tmall.com/item.htm?id=19464251895',
'https://detail.tmall.com/item.htm?id=554583889496',
'https://detail.tmall.com/item.htm?id=555673239683',
'https://detail.tmall.com/item.htm?id=521255569091',
'https://detail.tmall.com/item.htm?id=551315178985',
'https://detail.tmall.com/item.htm?id=542133102941',
'https://detail.tmall.hk/hk/item.htm?id=552360437632&skuId=3383711536314',
'https://detail.tmall.com/item.htm?id=520955843996',
'https://detail.tmall.com/item.htm?id=9514825867',
'https://detail.tmall.com/item.htm?id=521827178983'
]


filepath='./../test101.xlsx'
wb = openpyxl.load_workbook(filename=filepath)
ws = wb.get_sheet_by_name('Sheet')

row = 55
colum = 8





def get_peiliao(url):
    text = requests.get(url).text
    peiliao = re.compile(r'配料表：(.*?)</li>',re.S)
    peiliaos = re.findall(peiliao,text)
    return  (str(peiliaos))

def write_excel(row,data,url):
    ws.cell(row=row, column=8).value = data
    ws.cell(row=row, column=9).value = url


for url in LST:
    pei = get_peiliao(url)
    write_excel(row,pei,url)
    row  += 1

wb.save(filename='{}'.format(filepath))